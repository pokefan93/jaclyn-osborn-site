from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT = ROOT / "src" / "data" / "catalog-data.ts"

SHEET_ALIASES = {
    "Series": ["Series"],
    "Books": ["Books"],
    "Purchase": ["Purchase", "Availability & Notes", "Availability and Notes"],
    "Direct_Sale_Formats": ["Direct_Sale_Formats", "Direct Sales"],
    "Retailer_Links": ["Retailer_Links", "Retailer Links", "Store Links"],
}


def split_pipe(value: Any) -> list[str]:
    if value is None:
        return []
    return [part.strip() for part in str(value).split("|") if part.strip()]


def parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().lower() in {"1", "true", "yes", "y", "on", "checked"}


def normalize_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalize_availability_status(value: Any) -> str:
    normalized = normalize_text(value).lower().replace("-", " ").replace("_", " ")
    mapping = {
        "": "in_stock",
        "in stock": "in_stock",
        "available": "in_stock",
        "limited": "limited",
        "limited stock": "limited",
        "sold out": "sold_out",
        "preorder": "preorder",
        "pre order": "preorder",
    }
    return mapping.get(" ".join(normalized.split()), normalize_text(value) or "in_stock")


def normalize_purchase_mode(value: Any) -> str:
    normalized = normalize_text(value).lower().replace("-", " ").replace("_", " ")
    mapping = {
        "stripe payment link": "stripe_payment_link",
        "payment link": "stripe_payment_link",
        "direct checkout": "stripe_payment_link",
        "direct checkout stripe": "stripe_payment_link",
        "direct checkout (stripe)": "stripe_payment_link",
        "stripe buy button": "stripe_buy_button",
        "buy button": "stripe_buy_button",
        "embedded buy button": "stripe_buy_button",
        "unavailable": "unavailable",
        "not for sale right now": "unavailable",
        "sold out": "unavailable",
    }
    compact = " ".join(normalized.split())
    return mapping.get(compact, normalize_text(value))


def normalize_merchandising_flags(value: Any) -> list[str]:
    mapping = {
        "signed copy": "signed_copy",
        "direct from author": "direct_from_author",
        "new release": "new_release",
        "audiobook available": "audiobook_available",
        "preorder": "preorder",
    }
    flags: list[str] = []
    for part in split_pipe(value):
        normalized = part.lower().replace("-", " ").replace("_", " ")
        flags.append(mapping.get(" ".join(normalized.split()), part))
    return flags


def header_map(ws: Worksheet) -> dict[str, int]:
    return {str(cell.value): index for index, cell in enumerate(ws[1], start=1) if cell.value}


def get_sheet(workbook, canonical_name: str) -> Worksheet:
    aliases = SHEET_ALIASES[canonical_name]
    for alias in aliases:
        if alias in workbook.sheetnames:
            return workbook[alias]
    raise RuntimeError(
        f"Workbook is missing required sheet '{canonical_name}'. "
        f"Accepted names: {', '.join(aliases)}"
    )


def read_rows(ws: Worksheet) -> list[dict[str, Any]]:
    headers = header_map(ws)
    rows: list[dict[str, Any]] = []
    for row_index in range(2, ws.max_row + 1):
        if not any(ws.cell(row=row_index, column=col).value for col in headers.values()):
            continue
        row = {}
        for key, col in headers.items():
            row[key] = ws.cell(row=row_index, column=col).value
        rows.append(row)
    return rows


def to_float(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    return float(value)


def build_series(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    series_entries: list[dict[str, Any]] = []
    for row in rows:
        series_entries.append(
            {
                "slug": normalize_text(row.get("series_slug")),
                "name": normalize_text(row.get("name")),
                "lane": normalize_text(row.get("lane")),
                "tagline": normalize_text(row.get("tagline")),
                "description": normalize_text(row.get("description")),
                "status": normalize_text(row.get("status")) or "Series",
                "entryPoint": normalize_text(row.get("entry_point")),
                "gradient": [
                    normalize_text(row.get("gradient_start")),
                    normalize_text(row.get("gradient_mid")),
                    normalize_text(row.get("gradient_end")),
                ],
                "textColor": normalize_text(row.get("text_color")),
                "highlight": split_pipe(row.get("highlight")),
            }
        )
    return series_entries


def build_purchase_maps(
    purchase_rows: list[dict[str, Any]],
    direct_rows: list[dict[str, Any]],
    retailer_rows: list[dict[str, Any]],
) -> tuple[dict[str, dict[str, Any]], dict[str, list[dict[str, Any]]], dict[str, list[dict[str, Any]]]]:
    purchase_by_slug: dict[str, dict[str, Any]] = {}
    direct_by_slug: dict[str, list[dict[str, Any]]] = {}
    retailer_by_slug: dict[str, list[dict[str, Any]]] = {}

    for row in purchase_rows:
        slug = normalize_text(row.get("book_slug"))
        if not slug:
            continue
        purchase_by_slug[slug] = {
            "availabilityStatus": normalize_availability_status(row.get("availability_status")),
            "availabilityLabel": row.get("availability_label") or None,
            "priceNote": row.get("price_note") or None,
            "priceSnapshotDate": row.get("price_snapshot_date") or None,
            "merchandisingFlags": normalize_merchandising_flags(row.get("merchandising_flags")),
            "signedCopy": parse_bool(row.get("signed_copy")),
            "directFromAuthor": parse_bool(row.get("direct_from_author")),
            "signedCopyNote": row.get("signed_copy_note") or None,
            "shippingNote": row.get("shipping_note") or None,
            "fulfillmentNote": row.get("fulfillment_note") or None,
            "whereToBuyNote": row.get("where_to_buy_note") or None,
        }

    for row in direct_rows:
        slug = normalize_text(row.get("book_slug"))
        direct_sale_id = normalize_text(row.get("direct_sale_id"))
        fmt = normalize_text(row.get("format"))
        mode = normalize_text(row.get("purchase_mode"))
        if not slug or not direct_sale_id or not fmt or not mode:
            continue
        direct_by_slug.setdefault(slug, []).append(
            {
                "id": direct_sale_id,
                "label": normalize_text(row.get("label")) or fmt.title(),
                "format": fmt,
                "purchaseMode": normalize_purchase_mode(mode),
                "purchaseUrl": row.get("stripe_payment_link_url")
                or row.get("purchase_url")
                or None,
                "stripeBuyButtonId": row.get("stripe_buy_button_id") or None,
            }
        )

    for row in retailer_rows:
        slug = normalize_text(row.get("book_slug"))
        purchase_url = normalize_text(row.get("purchase_url"))
        link_id = normalize_text(row.get("link_id"))
        if not slug or not purchase_url or not link_id:
            continue
        retailer_by_slug.setdefault(slug, []).append(
            {
                "id": link_id,
                "retailer": normalize_text(row.get("retailer")) or "Retailer",
                "label": normalize_text(row.get("label")) or "Buy now",
                "format": normalize_text(row.get("format")) or None,
                "purchaseMode": "external_retailer",
                "purchaseUrl": purchase_url,
            }
        )

    return purchase_by_slug, direct_by_slug, retailer_by_slug


def build_books(
    book_rows: list[dict[str, Any]],
    purchase_by_slug: dict[str, dict[str, Any]],
    direct_by_slug: dict[str, list[dict[str, Any]]],
    retailer_by_slug: dict[str, list[dict[str, Any]]],
) -> list[dict[str, Any]]:
    books: list[dict[str, Any]] = []

    for row in book_rows:
        slug = normalize_text(row.get("book_slug"))
        if not slug:
            continue

        purchase = purchase_by_slug.get(
            slug,
            {
                "availabilityStatus": "in_stock",
                "availabilityLabel": None,
                "priceNote": None,
                "priceSnapshotDate": None,
                "merchandisingFlags": [],
                "signedCopy": False,
                "directFromAuthor": False,
                "signedCopyNote": None,
                "shippingNote": None,
                "fulfillmentNote": None,
                "whereToBuyNote": None,
            },
        )

        book = {
            "id": normalize_text(row.get("book_id")) or slug,
            "slug": slug,
            "title": normalize_text(row.get("title")),
            "seriesSlug": normalize_text(row.get("series_slug")),
            "seriesLabel": normalize_text(row.get("series_label")),
            "seriesOrder": to_float(row.get("series_order")),
            "releaseDate": normalize_text(row.get("release_date")),
            "shortHook": normalize_text(row.get("short_hook")),
            "description": normalize_text(row.get("description")),
            "genres": split_pipe(row.get("genres")),
            "vibes": split_pipe(row.get("vibes")),
            "tropes": split_pipe(row.get("tropes")),
            "formats": split_pipe(row.get("formats")),
            "featured": parse_bool(row.get("featured")),
            "coverPalette": normalize_text(row.get("cover_palette")) or "plum",
            "catalogStatus": normalize_text(row.get("catalog_status")),
            "purchase": {
                **purchase,
                "retailerLinks": retailer_by_slug.get(slug, []),
                "directSaleFormats": direct_by_slug.get(slug, []),
            },
        }
        books.append(book)

    return books


def format_ts_export(name: str, value: Any) -> str:
    return f"export const {name} = {json.dumps(value, indent=2, ensure_ascii=False)} as const;\n"


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    workbook = load_workbook(args.workbook, data_only=True)
    series_rows = read_rows(get_sheet(workbook, "Series"))
    book_rows = read_rows(get_sheet(workbook, "Books"))
    purchase_rows = read_rows(get_sheet(workbook, "Purchase"))
    direct_rows = read_rows(get_sheet(workbook, "Direct_Sale_Formats"))
    retailer_rows = read_rows(get_sheet(workbook, "Retailer_Links"))

    series = build_series(series_rows)
    purchase_by_slug, direct_by_slug, retailer_by_slug = build_purchase_maps(
        purchase_rows, direct_rows, retailer_rows
    )
    books = build_books(book_rows, purchase_by_slug, direct_by_slug, retailer_by_slug)

    header = (
        "// This file is generated by scripts/build_catalog_from_workbook.py\n"
        f"// Source workbook: {args.workbook.name}\n\n"
    )
    output = header + format_ts_export("generatedSeries", series) + "\n" + format_ts_export("generatedBooks", books)
    args.output.write_text(output, encoding="utf-8")
    print(args.output)


if __name__ == "__main__":
    main()
