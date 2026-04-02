from __future__ import annotations

import argparse
import json
import os
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.error import HTTPError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


STRIPE_API_BASE = "https://api.stripe.com/v1"
SHEET_ALIASES = {
    "Books": ["Books"],
    "Direct_Sale_Formats": ["Direct_Sale_Formats", "Direct Sales"],
}


class StripeApiError(RuntimeError):
    pass


def now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def normalize_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().lower() in {"1", "true", "yes", "y", "on", "checked"}


def bool_with_default(value: Any, default: bool) -> bool:
    if value is None or str(value).strip() == "":
        return default
    return normalize_bool(value)


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_purchase_mode(value: Any) -> str:
    normalized = normalize_text(value).lower().replace("-", " ").replace("_", " ")
    mapping = {
        "stripe payment link": "stripe_payment_link",
        "payment link": "stripe_payment_link",
        "direct checkout": "stripe_payment_link",
        "direct checkout stripe": "stripe_payment_link",
        "direct checkout (stripe)": "stripe_payment_link",
        "stripe buy button": "stripe_buy_button",
        "buy button": "stripe_buy_button",
        "embedded buy button": "stripe_buy_button",
        "unavailable": "unavailable",
        "not for sale right now": "unavailable",
        "sold out": "unavailable",
    }
    compact = " ".join(normalized.split())
    return mapping.get(compact, normalize_text(value))


def parse_country_list(raw: str) -> list[str]:
    normalized = raw.replace("|", ",").replace(";", ",")
    values = [item.strip().upper() for item in normalized.split(",")]
    return [value for value in values if value]


def header_map(ws: Worksheet) -> dict[str, int]:
    return {str(cell.value): index for index, cell in enumerate(ws[1], start=1) if cell.value}


def get_value(ws: Worksheet, row_index: int, headers: dict[str, int], key: str) -> Any:
    col = headers.get(key)
    return ws.cell(row=row_index, column=col).value if col else None


def set_value(ws: Worksheet, row_index: int, headers: dict[str, int], key: str, value: Any) -> None:
    col = headers.get(key)
    if not col:
        raise KeyError(f"Missing required column: {key}")
    ws.cell(row=row_index, column=col).value = value


def ensure_required_headers(ws: Worksheet, required: list[str]) -> dict[str, int]:
    headers = header_map(ws)
    missing = [key for key in required if key not in headers]
    if missing:
        raise RuntimeError(f"Worksheet '{ws.title}' is missing required columns: {', '.join(missing)}")
    return headers


def get_sheet(workbook, canonical_name: str) -> Worksheet:
    aliases = SHEET_ALIASES[canonical_name]
    for alias in aliases:
        if alias in workbook.sheetnames:
            return workbook[alias]
    raise RuntimeError(
        f"Workbook is missing required sheet '{canonical_name}'. "
        f"Accepted names: {', '.join(aliases)}"
    )


@dataclass
class StripeClient:
    secret_key: str | None
    dry_run: bool = False

    def request(self, method: str, path: str, form: list[tuple[str, str]] | None = None) -> dict[str, Any]:
        if self.dry_run:
            return {"id": "dry_run"}

        if not self.secret_key:
            raise StripeApiError("STRIPE_SECRET_KEY is required for live Stripe sync.")

        data = urlencode(form or []).encode("utf-8") if form is not None else None
        request = Request(
            f"{STRIPE_API_BASE}{path}",
            method=method,
            data=data,
            headers={
                "Authorization": f"Bearer {self.secret_key}",
                "Content-Type": "application/x-www-form-urlencoded",
            },
        )

        try:
            with urlopen(request, timeout=60) as response:
                return json.loads(response.read().decode("utf-8"))
        except HTTPError as exc:
            payload = exc.read().decode("utf-8", errors="replace")
            try:
                parsed = json.loads(payload)
                message = parsed.get("error", {}).get("message", payload)
            except json.JSONDecodeError:
                message = payload
            raise StripeApiError(f"Stripe API error {exc.code} for {path}: {message}") from exc

    def list_products(self) -> list[dict[str, Any]]:
        if self.dry_run:
            return []

        products: list[dict[str, Any]] = []
        starting_after = ""
        while True:
            form: list[tuple[str, str]] = [("limit", "100")]
            if starting_after:
                form.append(("starting_after", starting_after))
            page = self.request("GET", f"/products?{urlencode(form)}")
            data = page.get("data", [])
            products.extend(data)
            if not page.get("has_more") or not data:
                break
            starting_after = data[-1]["id"]
        return products

    def list_prices(self, product_id: str) -> list[dict[str, Any]]:
        if self.dry_run:
            return []

        prices: list[dict[str, Any]] = []
        starting_after = ""
        while True:
            form: list[tuple[str, str]] = [("limit", "100"), ("product", product_id)]
            if starting_after:
                form.append(("starting_after", starting_after))
            page = self.request("GET", f"/prices?{urlencode(form)}")
            data = page.get("data", [])
            prices.extend(data)
            if not page.get("has_more") or not data:
                break
            starting_after = data[-1]["id"]
        return prices

    def list_payment_links(self) -> list[dict[str, Any]]:
        if self.dry_run:
            return []

        links: list[dict[str, Any]] = []
        starting_after = ""
        while True:
            form: list[tuple[str, str]] = [("limit", "100")]
            if starting_after:
                form.append(("starting_after", starting_after))
            page = self.request("GET", f"/payment_links?{urlencode(form)}")
            data = page.get("data", [])
            links.extend(data)
            if not page.get("has_more") or not data:
                break
            starting_after = data[-1]["id"]
        return links

    def retrieve_price(self, price_id: str) -> dict[str, Any]:
        return self.request("GET", f"/prices/{price_id}")

    def create_product(
        self,
        *,
        name: str,
        description: str,
        metadata: dict[str, str],
        active: bool,
        image_url: str = "",
    ) -> dict[str, Any]:
        form = [
            ("name", name),
            ("description", description),
            ("active", "true" if active else "false"),
        ]
        if image_url:
            form.append(("images[0]", image_url))
        for key, value in metadata.items():
            if value:
                form.append((f"metadata[{key}]", value))
        return self.request("POST", "/products", form)

    def update_product(
        self,
        product_id: str,
        *,
        name: str,
        description: str,
        metadata: dict[str, str],
        active: bool,
        image_url: str = "",
    ) -> dict[str, Any]:
        form = [
            ("name", name),
            ("description", description),
            ("active", "true" if active else "false"),
        ]
        if image_url:
            form.append(("images[0]", image_url))
        for key, value in metadata.items():
            if value:
                form.append((f"metadata[{key}]", value))
        return self.request("POST", f"/products/{product_id}", form)

    def create_price(
        self,
        *,
        product_id: str,
        direct_sale_id: str,
        label: str,
        currency: str,
        unit_amount: int,
        metadata: dict[str, str],
    ) -> dict[str, Any]:
        form = [
            ("product", product_id),
            ("currency", currency.lower()),
            ("unit_amount", str(unit_amount)),
            ("nickname", label),
        ]
        for key, value in metadata.items():
            if value:
                form.append((f"metadata[{key}]", value))
        form.append((f"metadata[direct_sale_id]", direct_sale_id))
        return self.request("POST", "/prices", form)

    def archive_price(self, price_id: str) -> dict[str, Any]:
        return self.request("POST", f"/prices/{price_id}", [("active", "false")])

    def create_payment_link(
        self,
        *,
        price_id: str,
        quantity: int,
        allow_promotion_codes: bool,
        redirect_url: str,
        collect_shipping_address: bool,
        shipping_countries: list[str],
        metadata: dict[str, str],
    ) -> dict[str, Any]:
        form: list[tuple[str, str]] = [
            ("line_items[0][price]", price_id),
            ("line_items[0][quantity]", str(quantity)),
            ("allow_promotion_codes", "true" if allow_promotion_codes else "false"),
        ]
        if redirect_url:
            form.extend(
                [
                    ("after_completion[type]", "redirect"),
                    ("after_completion[redirect][url]", redirect_url),
                ]
            )
        if collect_shipping_address and shipping_countries:
            for index, country in enumerate(shipping_countries):
                form.append((f"shipping_address_collection[allowed_countries][{index}]", country))
        for key, value in metadata.items():
            if value:
                form.append((f"metadata[{key}]", value))
        return self.request("POST", "/payment_links", form)

    def deactivate_payment_link(self, payment_link_id: str) -> dict[str, Any]:
        return self.request("POST", f"/payment_links/{payment_link_id}", [("active", "false")])


def default_output_path(input_path: Path) -> Path:
    return input_path.with_name(f"{input_path.stem}_stripe_synced{input_path.suffix}")


def build_product_description(book_row: dict[str, str]) -> str:
    short_hook = normalize_text(book_row.get("short_hook"))
    description = normalize_text(book_row.get("description"))
    return short_hook or description[:5000]


def read_sheet_rows(ws: Worksheet, headers: dict[str, int]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row_index in range(2, ws.max_row + 1):
        if not any(ws.cell(row=row_index, column=col).value for col in headers.values()):
            continue
        row = {"_row_index": row_index}
        for key, col in headers.items():
            row[key] = ws.cell(row=row_index, column=col).value
        rows.append(row)
    return rows


def sync_products(
    client: StripeClient,
    books_ws: Worksheet,
    book_headers: dict[str, int],
    books: list[dict[str, Any]],
    cover_base_url: str = "",
) -> dict[str, str]:
    product_ids_by_slug: dict[str, str] = {}
    existing_by_slug: dict[str, dict[str, Any]] = {}

    if not client.dry_run:
        for product in client.list_products():
            metadata = product.get("metadata") or {}
            book_slug = metadata.get("book_slug")
            if book_slug and book_slug not in existing_by_slug:
                existing_by_slug[book_slug] = product

    for book in books:
        row_index = book["_row_index"]
        sync_enabled = bool_with_default(book.get("sync_product_to_stripe"), True)
        if not sync_enabled:
            continue

        book_slug = normalize_text(book.get("book_slug"))
        title = normalize_text(book.get("title"))
        product_id = normalize_text(book.get("stripe_product_id"))
        active = bool_with_default(book.get("stripe_product_active"), True)

        metadata = {
            "source": "jaclyn-site",
            "book_slug": book_slug,
            "series_slug": normalize_text(book.get("series_slug")),
            "catalog_status": normalize_text(book.get("catalog_status")),
        }
        description = build_product_description(book)
        image_url = (
            f"{cover_base_url.rstrip('/')}/{book_slug}.jpg" if cover_base_url else ""
        )

        existing = existing_by_slug.get(book_slug)
        if existing and not product_id:
            product_id = existing["id"]

        if client.dry_run:
            if product_id:
                print(f"[dry-run] update product {product_id} for {book_slug}")
            else:
                print(f"[dry-run] create product for {book_slug}")
                product_id = f"dry_run_prod_{book_slug}"
        else:
            if product_id:
                product = client.update_product(
                    product_id,
                    name=title,
                    description=description,
                    metadata=metadata,
                    active=active,
                    image_url=image_url,
                )
            elif existing:
                product = client.update_product(
                    existing["id"],
                    name=title,
                    description=description,
                    metadata=metadata,
                    active=active,
                    image_url=image_url,
                )
            else:
                product = client.create_product(
                    name=title,
                    description=description,
                    metadata=metadata,
                    active=active,
                    image_url=image_url,
                )
            product_id = product["id"]
            set_value(books_ws, row_index, book_headers, "stripe_product_id", product_id)
            set_value(books_ws, row_index, book_headers, "stripe_last_synced_at", now_iso())
            set_value(books_ws, row_index, book_headers, "stripe_sync_notes", "")

        product_ids_by_slug[book_slug] = product_id

    return product_ids_by_slug


def sync_direct_sale_formats(
    client: StripeClient,
    direct_ws: Worksheet,
    direct_headers: dict[str, int],
    direct_rows: list[dict[str, Any]],
    product_ids_by_slug: dict[str, str],
) -> None:
    existing_payment_links_by_sale_id: dict[str, dict[str, Any]] = {}
    existing_prices_by_product: dict[str, list[dict[str, Any]]] = {}

    if not client.dry_run:
        for link in client.list_payment_links():
            metadata = link.get("metadata") or {}
            direct_sale_id = metadata.get("direct_sale_id")
            if direct_sale_id and direct_sale_id not in existing_payment_links_by_sale_id:
                existing_payment_links_by_sale_id[direct_sale_id] = link

    for row in direct_rows:
        row_index = row["_row_index"]
        if not normalize_bool(row.get("sync_to_stripe")):
            continue

        book_slug = normalize_text(row.get("book_slug"))
        direct_sale_id = normalize_text(row.get("direct_sale_id"))
        label = normalize_text(row.get("label")) or normalize_text(row.get("book_title"))
        currency = normalize_text(row.get("currency") or "usd").lower()
        unit_amount_raw = normalize_text(row.get("unit_amount"))
        purchase_mode = normalize_purchase_mode(row.get("purchase_mode") or "stripe_payment_link")

        if purchase_mode not in {"stripe_payment_link", "stripe_buy_button"}:
            raise RuntimeError(
                f"Direct sale row {direct_sale_id or book_slug} must use stripe_payment_link "
                "or stripe_buy_button for Stripe sync."
            )

        if not unit_amount_raw:
            raise RuntimeError(
                f"Direct sale row {direct_sale_id or book_slug} is missing unit_amount."
            )

        product_id = normalize_text(row.get("stripe_product_id")) or product_ids_by_slug.get(book_slug, "")
        if not product_id:
            raise RuntimeError(
                f"Direct sale row {direct_sale_id or book_slug} has no Stripe product id."
            )

        try:
            unit_amount = int(float(unit_amount_raw))
        except ValueError as exc:
            raise RuntimeError(
                f"Direct sale row {direct_sale_id or book_slug} has invalid unit_amount '{unit_amount_raw}'. "
                "Use Stripe minor units, e.g. 1999 for $19.99."
            ) from exc

        metadata = {
            "source": "jaclyn-site",
            "book_slug": book_slug,
            "direct_sale_id": direct_sale_id,
            "format": normalize_text(row.get("format")),
        }

        existing_price_id = normalize_text(row.get("stripe_price_id"))
        price_id = existing_price_id
        create_new_price = True

        if not existing_price_id and not client.dry_run:
            product_prices = existing_prices_by_product.get(product_id)
            if product_prices is None:
                product_prices = client.list_prices(product_id)
                existing_prices_by_product[product_id] = product_prices
            for existing_price in product_prices:
                metadata_match = (
                    normalize_text((existing_price.get("metadata") or {}).get("direct_sale_id"))
                    == direct_sale_id
                )
                if (
                    metadata_match
                    and int(existing_price.get("unit_amount") or 0) == unit_amount
                    and normalize_text(existing_price.get("currency")).lower() == currency
                    and bool(existing_price.get("active"))
                ):
                    existing_price_id = existing_price["id"]
                    break

        if existing_price_id:
            if client.dry_run:
                print(f"[dry-run] inspect existing price {existing_price_id} for {direct_sale_id}")
                create_new_price = False
                price_id = existing_price_id
            else:
                current_price = client.retrieve_price(existing_price_id)
                if (
                    int(current_price.get("unit_amount") or 0) == unit_amount
                    and normalize_text(current_price.get("currency")).lower() == currency
                    and normalize_text(current_price.get("product")) == product_id
                    and bool(current_price.get("active"))
                ):
                    create_new_price = False
                    price_id = existing_price_id

        if create_new_price:
            if client.dry_run:
                price_id = f"dry_run_price_{direct_sale_id}"
                print(
                    f"[dry-run] create price for {direct_sale_id} "
                    f"({currency} {unit_amount}) on product {product_id}"
                )
            else:
                new_price = client.create_price(
                    product_id=product_id,
                    direct_sale_id=direct_sale_id,
                    label=label,
                    currency=currency,
                    unit_amount=unit_amount,
                    metadata=metadata,
                )
                price_id = new_price["id"]
                if existing_price_id and existing_price_id != price_id:
                    try:
                        client.archive_price(existing_price_id)
                    except StripeApiError:
                        pass
                set_value(direct_ws, row_index, direct_headers, "stripe_price_id", price_id)

        existing_payment_link_id = normalize_text(row.get("stripe_payment_link_id"))
        existing_payment_link_url = normalize_text(row.get("stripe_payment_link_url"))
        purchase_url = normalize_text(row.get("purchase_url"))
        redirect_url = normalize_text(row.get("after_completion_redirect_url"))
        allow_promotion_codes = normalize_bool(row.get("allow_promotion_codes"))
        collect_shipping_address = normalize_bool(row.get("collect_shipping_address"))
        shipping_countries = parse_country_list(normalize_text(row.get("shipping_countries")))

        payment_link_id = existing_payment_link_id
        payment_link_url = existing_payment_link_url or purchase_url

        if not existing_payment_link_id and not client.dry_run:
            existing_link = existing_payment_links_by_sale_id.get(direct_sale_id)
            if existing_link and not create_new_price:
                existing_payment_link_id = existing_link["id"]
                existing_payment_link_url = existing_link.get("url") or existing_payment_link_url
                payment_link_id = existing_payment_link_id
                payment_link_url = existing_payment_link_url or purchase_url

        needs_new_payment_link = not existing_payment_link_id or create_new_price

        if needs_new_payment_link:
            if client.dry_run:
                payment_link_id = f"dry_run_plink_{direct_sale_id}"
                payment_link_url = f"https://buy.stripe.com/{payment_link_id}"
                print(f"[dry-run] create payment link for {direct_sale_id} using price {price_id}")
            else:
                payment_link = client.create_payment_link(
                    price_id=price_id,
                    quantity=1,
                    allow_promotion_codes=allow_promotion_codes,
                    redirect_url=redirect_url,
                    collect_shipping_address=collect_shipping_address,
                    shipping_countries=shipping_countries,
                    metadata=metadata,
                )
                payment_link_id = payment_link["id"]
                payment_link_url = payment_link["url"]
                if existing_payment_link_id and existing_payment_link_id != payment_link_id:
                    try:
                        client.deactivate_payment_link(existing_payment_link_id)
                    except StripeApiError:
                        pass

                set_value(direct_ws, row_index, direct_headers, "stripe_payment_link_id", payment_link_id)
                set_value(direct_ws, row_index, direct_headers, "stripe_payment_link_url", payment_link_url)
                set_value(direct_ws, row_index, direct_headers, "purchase_url", payment_link_url)
        elif existing_payment_link_id:
            set_value(direct_ws, row_index, direct_headers, "stripe_payment_link_id", existing_payment_link_id)
            set_value(direct_ws, row_index, direct_headers, "stripe_payment_link_url", existing_payment_link_url)
            set_value(direct_ws, row_index, direct_headers, "purchase_url", existing_payment_link_url)

        if not client.dry_run:
            set_value(direct_ws, row_index, direct_headers, "stripe_product_id", product_id)
            set_value(direct_ws, row_index, direct_headers, "stripe_price_id", price_id)
            set_value(direct_ws, row_index, direct_headers, "stripe_last_synced_at", now_iso())
            set_value(direct_ws, row_index, direct_headers, "stripe_sync_notes", "")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--output", type=Path)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--cover-base-url", default="")
    args = parser.parse_args()

    workbook_path = args.workbook
    output_path = args.output or default_output_path(workbook_path)

    wb = load_workbook(workbook_path)
    books_ws = get_sheet(wb, "Books")
    direct_ws = get_sheet(wb, "Direct_Sale_Formats")

    book_headers = ensure_required_headers(
        books_ws,
        [
            "book_slug",
            "title",
            "short_hook",
            "description",
            "series_slug",
            "catalog_status",
            "sync_product_to_stripe",
            "stripe_product_id",
            "stripe_product_active",
            "stripe_last_synced_at",
            "stripe_sync_notes",
        ],
    )
    direct_headers = ensure_required_headers(
        direct_ws,
        [
            "sync_to_stripe",
            "book_slug",
            "book_title",
            "direct_sale_id",
            "label",
            "format",
            "purchase_mode",
            "unit_amount",
            "currency",
            "collect_shipping_address",
            "shipping_countries",
            "allow_promotion_codes",
            "after_completion_redirect_url",
            "purchase_url",
            "stripe_product_id",
            "stripe_price_id",
            "stripe_payment_link_id",
            "stripe_payment_link_url",
            "stripe_buy_button_id",
            "stripe_last_synced_at",
            "stripe_sync_notes",
        ],
    )

    books = read_sheet_rows(books_ws, book_headers)
    direct_rows = read_sheet_rows(direct_ws, direct_headers)

    client = StripeClient(
        secret_key=os.environ.get("STRIPE_SECRET_KEY"),
        dry_run=args.dry_run,
    )

    product_ids_by_slug = sync_products(
        client,
        books_ws,
        book_headers,
        books,
        cover_base_url=normalize_text(args.cover_base_url),
    )
    sync_direct_sale_formats(client, direct_ws, direct_headers, direct_rows, product_ids_by_slug)

    if args.dry_run:
        print("Dry run complete. No Stripe or workbook changes were saved.")
        return

    wb.save(output_path)
    print(output_path)


if __name__ == "__main__":
    main()
