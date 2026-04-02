from __future__ import annotations

import argparse
import json
import re
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


ROOT = Path(__file__).resolve().parents[1]
CATALOG_DATA = ROOT / "src" / "data" / "catalog-data.ts"
DEFAULT_OUTPUT = (
    Path.home()
    / "Downloads"
    / f"jaclyn_osborn_catalog_google_sheets_ready_{date.today().isoformat()}.xlsx"
)

HEADER_FILL = PatternFill("solid", fgColor="241920")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FILL = PatternFill("solid", fgColor="F3E4DC")
SECTION_FONT = Font(color="241920", bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D8C5BF"),
    right=Side(style="thin", color="D8C5BF"),
    top=Side(style="thin", color="D8C5BF"),
    bottom=Side(style="thin", color="D8C5BF"),
)


def load_catalog() -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    text = CATALOG_DATA.read_text(encoding="utf-8")
    series_match = re.search(
        r"export const generatedSeries = (\[.*?\]) as const;\n\nexport const generatedBooks =",
        text,
        re.S,
    )
    books_match = re.search(
        r"export const generatedBooks = (\[.*\]) as const;\s*$",
        text,
        re.S,
    )

    if not series_match or not books_match:
        raise RuntimeError("Could not parse catalog-data.ts exports.")

    series = json.loads(series_match.group(1))
    books = json.loads(books_match.group(1))
    return series, books


def pipe_join(values: list[str]) -> str:
    return " | ".join(value for value in values if value)


def style_header(ws: Worksheet, header_row: int = 1) -> None:
    for cell in ws[header_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = ws.dimensions


def style_body(ws: Worksheet, start_row: int = 2) -> None:
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER


def autosize_columns(ws: Worksheet, max_width: int = 56) -> None:
    for column_cells in ws.columns:
        length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            length = max(length, len(value))
        adjusted = min(max(length + 2, 10), max_width)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted


def make_admin_docs_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Admin_Docs")
    ws.append(["Section", "Details"])

    rows = [
        (
            "What this workbook is",
            (
                "Google-Sheets-ready catalog workbook based on the current Jaclyn Osborn "
                "site data. Upload this .xlsx to Google Drive and open it with Google Sheets."
            ),
        ),
        (
            "Current site source of truth",
            (
                "This export was generated from src/data/catalog-data.ts, which is what the "
                "site currently uses for books, series, purchase metadata, and retailer links."
            ),
        ),
        (
            "Tab overview",
            (
                "Series = one row per series. Books = one row per book. Purchase = one row per "
                "book-level purchase state. Direct_Sale_Formats = one row per direct-sale edition. "
                "Retailer_Links = one row per retailer link. Reference_Lists = allowed values."
            ),
        ),
        (
            "Do not change stable keys",
            (
                "Once a book or series is live on the site, do not casually rename book_id, "
                "book_slug, or series_slug. Those should stay stable."
            ),
        ),
        (
            "Multi-value fields",
            (
                "Use a pipe separator for list fields in Google Sheets: genre1 | genre2 | genre3. "
                "This workbook exports those fields in that format already."
            ),
        ),
        (
            "How to add a new book",
            (
                "1. Add or confirm the series row in Series if needed. 2. Add a row in Books. "
                "3. Add a matching row in Purchase. 4. Add retailer rows in Retailer_Links. "
                "5. Add direct-sale rows in Direct_Sale_Formats only if Jaclyn is selling that "
                "edition directly."
            ),
        ),
        (
            "How to update a description or series order",
            (
                "Edit the row in Books. Update title, short_hook, description, release_date, "
                "series_slug, or series_order there."
            ),
        ),
        (
            "How to handle pricing",
            (
                "For GitHub Pages, Stripe or the retailer is the checkout source of truth. "
                "price_note is editorial only. If you show price text on the site, update the "
                "sheet when the live Stripe or retailer price changes."
            ),
        ),
        (
            "Stripe direct sales",
            (
                "Use Stripe Payment Links as the default direct-sale method. Create the product "
                "and price in Stripe Dashboard, create a Payment Link, and paste the Payment "
                "Link URL into Direct_Sale_Formats only if you are managing links manually. "
                "If you use the Stripe sync workflow, leave purchase_url blank and the sync "
                "script will create the link for you."
            ),
        ),
        (
            "Bulk Stripe sync script",
            (
                "Use scripts/sync_stripe_from_workbook.py to create or update Stripe products in "
                "bulk from this workbook. The script uses STRIPE_SECRET_KEY from your local "
                "environment and must never be run from the GitHub Pages frontend."
            ),
        ),
        (
            "Stripe sync workflow",
            (
                "1. Fill or review Books and Direct_Sale_Formats. 2. Set sync_product_to_stripe "
                "or sync_to_stripe to TRUE where needed. 3. Run the sync script locally. 4. The "
                "script writes Stripe product, price, and payment-link IDs back into the workbook. "
                "Leave after_completion_redirect_url blank to use the site thank-you page by "
                "default, or fill it only when a row needs a custom redirect."
            ),
        ),
        (
            "Which direct-sale mode to use",
            (
                "Preferred: stripe_payment_link. Optional only: stripe_buy_button. Do not use "
                "Stripe secret keys, custom checkout sessions, or a custom cart on GitHub Pages."
            ),
        ),
        (
            "When to use retailer links",
            (
                "Use Retailer_Links for Kindle, paperback, audiobook, Kobo, Bookshop, Libro.fm, "
                "or any edition fulfilled somewhere else."
            ),
        ),
        (
            "Availability values",
            (
                "in_stock, limited, sold_out, and preorder are editorial labels only. They do "
                "not represent real-time stock."
            ),
        ),
        (
            "How to mark sold out",
            (
                "In the Purchase tab, set availability_status to sold_out. If only the direct "
                "edition is sold out but retailer editions should remain available, keep the "
                "Retailer_Links rows and change the matching Direct_Sale_Formats row to "
                "purchase_mode = unavailable or leave sync_to_stripe off."
            ),
        ),
        (
            "Audiobook rows",
            (
                "Audiobook can appear in Books.formats and can also have a Direct_Sale_Formats "
                "row if you ever want direct audio delivery. If you only want retailer audio, "
                "leave the audiobook direct-sale row disabled and use Retailer_Links."
            ),
        ),
        (
            "When to mark sold out",
            (
                "Mark sold_out when a direct signed or special edition should no longer be "
                "purchased from the site, even if retailer editions remain available."
            ),
        ),
        (
            "Merchandising flags",
            (
                "Available values: signed_copy, direct_from_author, new_release, "
                "audiobook_available, preorder."
            ),
        ),
        (
            "Current direct-sale state",
            (
                "The current site export has retailer links populated, but no active direct-sale "
                "rows yet. The Direct_Sale_Formats tab is ready for Stripe links whenever needed."
            ),
        ),
    ]

    for row in rows:
        ws.append(list(row))

    style_header(ws)
    style_body(ws)
    autosize_columns(ws, max_width=100)


def make_summary_sheet(
    wb: Workbook, series: list[dict[str, Any]], books: list[dict[str, Any]]
) -> None:
    ws = wb.create_sheet("Summary")
    ws.append(["Metric", "Value"])

    retailer_link_count = sum(len(book["purchase"]["retailerLinks"]) for book in books)
    direct_sale_count = sum(len(book["purchase"]["directSaleFormats"]) for book in books)
    seeded_direct_sale_candidates = sum(
        1 for book in books for fmt in book.get("formats", []) if fmt != "audiobook"
    )
    sold_out_count = sum(
        1 for book in books if book["purchase"].get("availabilityStatus") == "sold_out"
    )
    featured_titles = pipe_join([book["title"] for book in books if book.get("featured")])

    rows = [
        ("Export date", date.today().isoformat()),
        ("Series rows", len(series)),
        ("Book rows", len(books)),
        ("Retailer link rows", retailer_link_count),
        ("Direct-sale rows", direct_sale_count),
        ("Seeded direct-sale candidate rows", seeded_direct_sale_candidates),
        ("Sold-out titles", sold_out_count),
        ("Featured titles", featured_titles),
        ("Catalog source file", str(CATALOG_DATA.relative_to(ROOT))),
    ]

    for row in rows:
        ws.append(list(row))

    style_header(ws)
    style_body(ws)
    autosize_columns(ws, max_width=80)


def make_reference_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Reference_Lists")
    ws.append(["Category", "Value", "Notes"])

    rows = [
        ("format", "ebook", "Digital edition"),
        ("format", "paperback", "Print paperback edition"),
        ("format", "hardcover", "Print hardcover edition"),
        ("format", "audiobook", "Audio edition"),
        ("availability_status", "in_stock", "Editorial label only"),
        ("availability_status", "limited", "Editorial label only"),
        ("availability_status", "sold_out", "Editorial label only"),
        ("availability_status", "preorder", "Editorial label only"),
        ("purchase_mode", "external_retailer", "Retailer-hosted purchase link"),
        ("purchase_mode", "stripe_payment_link", "Recommended direct-sale mode"),
        ("purchase_mode", "stripe_buy_button", "Optional embedded Stripe mode"),
        ("purchase_mode", "unavailable", "No purchase available"),
        ("merchandising_flag", "signed_copy", "Signed edition or signed copy messaging"),
        ("merchandising_flag", "direct_from_author", "Direct sale fulfilled by author"),
        ("merchandising_flag", "new_release", "Highlight as a new release"),
        ("merchandising_flag", "audiobook_available", "Audiobook exists"),
        ("merchandising_flag", "preorder", "Title or edition is preorder"),
    ]

    for row in rows:
        ws.append(list(row))

    style_header(ws)
    style_body(ws)
    autosize_columns(ws)


def make_series_sheet(wb: Workbook, series: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Series")
    ws.append(
        [
            "series_slug",
            "name",
            "lane",
            "tagline",
            "description",
            "status",
            "entry_point",
            "gradient_start",
            "gradient_mid",
            "gradient_end",
            "text_color",
            "highlight",
        ]
    )

    for item in series:
        ws.append(
            [
                item["slug"],
                item["name"],
                item["lane"],
                item["tagline"],
                item["description"],
                item["status"],
                item["entryPoint"],
                item["gradient"][0] if item.get("gradient") else "",
                item["gradient"][1] if item.get("gradient") else "",
                item["gradient"][2] if item.get("gradient") else "",
                item.get("textColor", ""),
                pipe_join(item.get("highlight", [])),
            ]
        )

    style_header(ws)
    style_body(ws)
    autosize_columns(ws, max_width=72)


def make_books_sheet(wb: Workbook, books: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Books")
    ws.append(
        [
            "book_id",
            "book_slug",
            "title",
            "series_slug",
            "series_label",
            "series_order",
            "release_date",
            "short_hook",
            "description",
            "genres",
            "vibes",
            "tropes",
            "formats",
            "featured",
            "cover_palette",
            "catalog_status",
            "sync_product_to_stripe",
            "stripe_product_id",
            "stripe_product_active",
            "stripe_last_synced_at",
            "stripe_sync_notes",
        ]
    )

    for book in books:
        ws.append(
            [
                book["id"],
                book["slug"],
                book["title"],
                book.get("seriesSlug", ""),
                book.get("seriesLabel", ""),
                book.get("seriesOrder", ""),
                book.get("releaseDate", ""),
                book.get("shortHook", ""),
                book.get("description", ""),
                pipe_join(book.get("genres", [])),
                pipe_join(book.get("vibes", [])),
                pipe_join(book.get("tropes", [])),
                pipe_join(book.get("formats", [])),
                bool(book.get("featured")),
                book.get("coverPalette", ""),
                book.get("catalogStatus", ""),
                True,
                "",
                True,
                "",
                "",
            ]
        )

    style_header(ws)
    style_body(ws)
    autosize_columns(ws, max_width=96)


def make_purchase_sheet(wb: Workbook, books: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Purchase")
    ws.append(
        [
            "book_slug",
            "book_title",
            "availability_status",
            "availability_label",
            "price_note",
            "price_snapshot_date",
            "merchandising_flags",
            "signed_copy",
            "direct_from_author",
            "signed_copy_note",
            "shipping_note",
            "fulfillment_note",
            "where_to_buy_note",
        ]
    )

    for book in books:
        purchase = book["purchase"]
        ws.append(
            [
                book["slug"],
                book["title"],
                purchase.get("availabilityStatus", ""),
                purchase.get("availabilityLabel") or "",
                purchase.get("priceNote") or "",
                purchase.get("priceSnapshotDate") or "",
                pipe_join(purchase.get("merchandisingFlags", [])),
                bool(purchase.get("signedCopy")),
                bool(purchase.get("directFromAuthor")),
                purchase.get("signedCopyNote") or "",
                purchase.get("shippingNote") or "",
                purchase.get("fulfillmentNote") or "",
                purchase.get("whereToBuyNote") or "",
            ]
        )

    style_header(ws)
    style_body(ws)
    autosize_columns(ws, max_width=90)


def make_direct_sale_sheet(wb: Workbook, books: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Direct_Sale_Formats")
    ws.append(
        [
            "sync_to_stripe",
            "book_slug",
            "book_title",
            "direct_sale_id",
            "label",
            "format",
            "purchase_mode",
            "unit_amount",
            "currency",
            "collect_shipping_address",
            "shipping_countries",
            "allow_promotion_codes",
            "after_completion_redirect_url",
            "purchase_url",
            "stripe_product_id",
            "stripe_price_id",
            "stripe_payment_link_id",
            "stripe_payment_link_url",
            "stripe_buy_button_id",
            "stripe_last_synced_at",
            "stripe_sync_notes",
        ]
    )

    rows_added = 0
    for book in books:
        existing_rows = book["purchase"].get("directSaleFormats", [])
        if existing_rows:
            for item in existing_rows:
                is_physical = item.get("format") in {"paperback", "hardcover"}
                ws.append(
                    [
                        item.get("purchaseMode") in {"stripe_payment_link", "stripe_buy_button"},
                        book["slug"],
                        book["title"],
                        item.get("id", ""),
                        item.get("label", ""),
                        item.get("format", ""),
                        item.get("purchaseMode", ""),
                        "",
                        "usd",
                        is_physical,
                        "US" if is_physical else "",
                        True,
                        "",
                        item.get("purchaseUrl") or "",
                        "",
                        "",
                        "",
                        "",
                        item.get("stripeBuyButtonId") or "",
                        "",
                        "",
                    ]
                )
                rows_added += 1
            continue

        for fmt in book.get("formats", []):
            is_physical = fmt in {"paperback", "hardcover"}
            label = {
                "ebook": "eBook",
                "paperback": "Paperback",
                "hardcover": "Hardcover",
                "audiobook": "Audiobook",
            }.get(fmt, fmt.title())
            ws.append(
                [
                    False,
                    book["slug"],
                    book["title"],
                    f"{book['slug']}-{fmt}",
                    label,
                    fmt,
                    "stripe_payment_link",
                    "",
                    "usd",
                    is_physical,
                    "US" if is_physical else "",
                    True,
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                ]
            )
            rows_added += 1

    style_header(ws)
    if ws.max_row > 1:
        style_body(ws)
    autosize_columns(ws, max_width=84)


def make_retailer_sheet(wb: Workbook, books: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Retailer_Links")
    ws.append(
        [
            "book_slug",
            "book_title",
            "link_id",
            "retailer",
            "label",
            "format",
            "purchase_mode",
            "purchase_url",
        ]
    )

    for book in books:
        for item in book["purchase"].get("retailerLinks", []):
            ws.append(
                [
                    book["slug"],
                    book["title"],
                    item.get("id", ""),
                    item.get("retailer", ""),
                    item.get("label", ""),
                    item.get("format") or "",
                    item.get("purchaseMode", ""),
                    item.get("purchaseUrl") or "",
                ]
            )

    style_header(ws)
    if ws.max_row > 1:
        style_body(ws)
    autosize_columns(ws, max_width=100)


def build_workbook(output_path: Path) -> Path:
    series, books = load_catalog()

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    make_admin_docs_sheet(wb)
    make_summary_sheet(wb, series, books)
    make_reference_sheet(wb)
    make_series_sheet(wb, series)
    make_books_sheet(wb, books)
    make_purchase_sheet(wb, books)
    make_direct_sale_sheet(wb, books)
    make_retailer_sheet(wb, books)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    output_path = build_workbook(args.output)
    print(output_path)


if __name__ == "__main__":
    main()
